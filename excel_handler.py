# excel_handler.py
import os
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from tkinter import filedialog, messagebox
from path_utils import get_resource_path, ensure_dir

FORMULA_FILE = get_resource_path("formulas.json")
EXPORT_FOLDER = get_resource_path("exports")


def load_formulas():
    """Load formulas from formulas.json."""
    if os.path.exists(FORMULA_FILE):
        with open(FORMULA_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def ensure_export_folder():
    """Ensure the export folder exists."""
    if not os.path.exists(EXPORT_FOLDER):
        os.makedirs(EXPORT_FOLDER)


def export_to_excel(columns, data, ask_filename=True):
    """
    Export TreeView data to Excel with formulas and formatting.
    :param columns: list of column names
    :param data: list of row values (from TreeView)
    :param ask_filename: if True, open Save As dialog
    """
    ensure_export_folder()
    formulas = load_formulas()

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    # --- Write Headers ---
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # --- Write Data ---
    for row_idx, row_values in enumerate(data, start=2):
        for col_idx, value in enumerate(row_values, start=1):
            col_name = columns[col_idx - 1]

            if col_name in formulas and formulas[col_name].get("formula"):
                # Write formula using Excel column letters
                excel_formula = convert_formula(formulas[col_name]["formula"], columns, row_idx)
                cell = ws.cell(row=row_idx, column=col_idx)
                if value == "📝 Formula":  # If it's our formula placeholder
                    cell.value = excel_formula
                    if excel_formula.startswith("="):
                        cell.number_format = '#,##0.00'
                else:
                    cell.value = value  # Keep the actual value
            else:
                cell = ws.cell(row=row_idx, column=col_idx)
                try:
                    # Try to convert numeric strings to actual numbers
                    if isinstance(value, str) and value.replace(".", "").replace(",", "").isdigit():
                        cell.value = float(value.replace(",", ""))
                        cell.number_format = '#,##0.00'
                    else:
                        cell.value = value
                except (ValueError, AttributeError):
                    cell.value = value

    # --- Adjust column widths ---
    for col_idx, col_name in enumerate(columns, start=1):
        excel_col = get_excel_column_name(col_idx)
        ws.column_dimensions[excel_col].width = max(15, len(col_name) + 2)

    # --- Choose filename ---
    filepath = None
    if ask_filename:
        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            initialdir=EXPORT_FOLDER,
            title="Save Excel File"
        )

    if not filepath:  # if dialog cancelled
        filename = datetime.now().strftime("%Y%m%d_%H%M%S.xlsx")
        filepath = os.path.join(EXPORT_FOLDER, filename)
    print(f"Row {row_idx}, Col {col_name}: {excel_formula if col_name in formulas else value}")

    wb.save(filepath)

    # Success popup
    messagebox.showinfo("Export Successful", f"Data exported to:\n{filepath}")
    return filepath


def convert_formula(named_formula, columns, row_idx):
    """
    Convert a column-name formula (e.g. =[Bill Wt (Qtl)] * [Basic Rate as per Bill])
    into an Excel-style formula with cell references.
    """
    if not named_formula:
        return ""

    # Special case for MRN No.
    if named_formula.strip().upper() == "=ROW()-1":
        return "=ROW()-1"

    # Start with the original formula
    excel_formula = named_formula

    # Replace [Column Name] with Excel cell references
    for col_name in columns:
        if f"[{col_name}]" in excel_formula:
            col_idx = columns.index(col_name) + 1  # 1-based index
            excel_col = get_excel_column_name(col_idx)
            excel_formula = excel_formula.replace(f"[{col_name}]", f"{excel_col}{row_idx}")

    # Ensure it starts with "="
    if not excel_formula.strip().startswith("="):
        excel_formula = "=" + excel_formula

    # Replace ROUND function if present
    if "ROUND(" in excel_formula:
        excel_formula = excel_formula.replace("ROUND(", "ROUND(")

    return excel_formula



def get_excel_column_name(n):
    """
    Convert a 1-based column index into Excel-style column letters.
    Example: 1 -> A, 27 -> AA
    """
    result = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        result = chr(65 + remainder) + result
    return result
